#!/usr/bin/env python3
"""Export CSVs to PDF (reportlab) and XLSX (openpyxl)."""

import csv
import os
import sys
from datetime import datetime
from pathlib import Path

BASE = Path(os.getcwd())
OUT = BASE / "output"

ORG_CSV = BASE / "organizations.csv"
CONTACT_CSV = BASE / "contacts.csv"
PRODUCT_CSV = BASE / "products.csv"


def _ensure_deps():
    missing = []
    try:
        import reportlab
    except ImportError:
        missing.append("reportlab")
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
    if missing:
        import subprocess
        req = Path(__file__).parent.parent / "requirements.txt"
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", str(req)])


def load_csv(path):
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_org_index(orgs):
    return {r["id"]: r for r in orgs}


def build_prod_index(products):
    return {r["id"]: r.get("nombre", "") for r in products}


def resolve_products(pids_str, prod_index):
    if not pids_str:
        return []
    names = []
    for pid in pids_str.split(","):
        pid = pid.strip()
        if pid and pid in prod_index:
            names.append(prod_index[pid])
    return names


# ── Organizaciones PDF ──

def _pdf_orgs(orgs, ts):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle

    s_cell = ParagraphStyle("c", fontSize=6, leading=7, fontName="Helvetica")
    s_hdr = ParagraphStyle("h", fontSize=6, leading=7, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    s_title = ParagraphStyle("t", fontSize=10, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    blue = HexColor("#2F5496")
    gray = HexColor("#F2F2F2")
    border_c = HexColor("#BBBBBB")

    filepath = OUT / "organizations.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4),
                            leftMargin=3 * mm, rightMargin=3 * mm, topMargin=5 * mm, bottomMargin=5 * mm)
    story = []

    title_t = Table([[Paragraph(f"Organizations — {len(orgs)} records — {ts}", s_title)]], colWidths=[291 * mm])
    title_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), blue),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3), ("LEFTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(title_t)

    headers = ["#", "Name", "Industry", "Country", "Email", "Phone", "Website"]
    hdr = [Paragraph(h, s_hdr) for h in headers]
    data = [hdr]
    for o in orgs:
        data.append([
            Paragraph(o.get("num", ""), s_cell),
            Paragraph(o.get("nombre", "") or "-", s_cell),
            Paragraph(o.get("industria", "") or "-", s_cell),
            Paragraph(o.get("pais", "") or "-", s_cell),
            Paragraph(o.get("email", "") or "-", s_cell),
            Paragraph(o.get("telefono", "") or "-", s_cell),
            Paragraph(o.get("website", "") or "-", s_cell),
        ])
    col_w = [6 * mm, 30 * mm, 22 * mm, 16 * mm, 30 * mm, 22 * mm, 165 * mm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, border_c),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [None, gray]),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    doc.build(story)
    print(f"  PDF: {filepath}")


# ── Contactos PDF ──

def _pdf_contacts(contacts, org_idx, prod_idx, ts):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle

    s_cell = ParagraphStyle("c", fontSize=6, leading=7, fontName="Helvetica")
    s_hdr = ParagraphStyle("h", fontSize=6, leading=7, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    s_title = ParagraphStyle("t", fontSize=10, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    blue = HexColor("#2F5496")
    gray = HexColor("#F2F2F2")
    border_c = HexColor("#BBBBBB")

    filepath = OUT / "contacts.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4),
                            leftMargin=3 * mm, rightMargin=3 * mm, topMargin=5 * mm, bottomMargin=5 * mm)
    story = []

    title_t = Table([[Paragraph(f"Contacts — {len(contacts)} records — {ts}", s_title)]], colWidths=[291 * mm])
    title_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), blue),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3), ("LEFTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(title_t)

    headers = ["#", "Name", "Role", "Email", "Phone", "Organization", "Origin", "City", "Products"]
    hdr = [Paragraph(h, s_hdr) for h in headers]
    data = [hdr]
    for c in contacts:
        oid = c.get("organization_id", "")
        org_name = org_idx[oid].get("nombre", "") if oid in org_idx else ""
        city = org_idx[oid].get("ciudad", "") if oid in org_idx else ""
        prods = resolve_products(c.get("product_ids", ""), prod_idx)
        prod_str = ", ".join(prods[:2])
        if len(prods) > 2:
            prod_str += f"...({len(prods)})"
        data.append([
            Paragraph(c.get("num", ""), s_cell),
            Paragraph(f"{c.get('nombre','') or '(org)'} {c.get('apellido','')}".strip(), s_cell),
            Paragraph(c.get("cargo", "") or "-", s_cell),
            Paragraph(c.get("email", "") or "-", s_cell),
            Paragraph(c.get("telefono_movil", "") or "-", s_cell),
            Paragraph(org_name or "-", s_cell),
            Paragraph(c.get("origen", "") or "-", s_cell),
            Paragraph(city or "-", s_cell),
            Paragraph(prod_str or "-", s_cell),
        ])
    col_w = [6 * mm, 22 * mm, 18 * mm, 30 * mm, 22 * mm, 30 * mm, 16 * mm, 14 * mm, 133 * mm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, border_c),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [None, gray]),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    doc.build(story)
    print(f"  PDF: {filepath}")


# ── Productos PDF ──

def _pdf_products(products, ts):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle

    s_cell = ParagraphStyle("c", fontSize=6, leading=7, fontName="Helvetica")
    s_hdr = ParagraphStyle("h", fontSize=6, leading=7, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    s_title = ParagraphStyle("t", fontSize=10, fontName="Helvetica-Bold", textColor=HexColor("#FFFFFF"))
    blue = HexColor("#2F5496")
    gray = HexColor("#F2F2F2")
    border_c = HexColor("#BBBBBB")

    filepath = OUT / "products.pdf"
    doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4),
                            leftMargin=3 * mm, rightMargin=3 * mm, topMargin=5 * mm, bottomMargin=5 * mm)
    story = []

    title_t = Table([[Paragraph(f"Products — {len(products)} records — {ts}", s_title)]], colWidths=[291 * mm])
    title_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), blue),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3), ("LEFTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(title_t)

    headers = ["#", "Name", "Category", "Presentation", "Volume", "Price Ref.", "Coverage", "Fleet", "Regulatory", "Notes"]
    hdr = [Paragraph(h, s_hdr) for h in headers]
    data = [hdr]
    for p in products:
        data.append([
            Paragraph(p.get("num", ""), s_cell),
            Paragraph(p.get("nombre", ""), s_cell),
            Paragraph(p.get("categoria", "") or "-", s_cell),
            Paragraph(p.get("presentacion", "") or "-", s_cell),
            Paragraph(p.get("volumen", "") or "-", s_cell),
            Paragraph(p.get("precio_ref", "") or "-", s_cell),
            Paragraph(p.get("cobertura", "") or "-", s_cell),
            Paragraph(p.get("flota", "") or "-", s_cell),
            Paragraph(p.get("regulatorio", "") or "-", s_cell),
            Paragraph(p.get("notas", "") or "-", s_cell),
        ])
    col_w = [6 * mm, 30 * mm, 22 * mm, 24 * mm, 20 * mm, 18 * mm, 20 * mm, 10 * mm, 22 * mm, 119 * mm]
    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, border_c),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [None, gray]),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    doc.build(story)
    print(f"  PDF: {filepath}")


# ── XLSX (todas) ──

def _xlsx_all(orgs, contacts, org_idx, prod_idx, products, ts):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_font = Font(name="Arial", size=9)

    def write_sheet(ws, headers, rows, widths):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border
        for ri, row in enumerate(rows, 2):
            for ci, v in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = cell_font
                cell.border = border
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    OUT.mkdir(parents=True, exist_ok=True)

    # Organizations
    wb = Workbook()
    ws = wb.active
    ws.title = "Organizations"
    h_org = ["#", "Name", "Industry", "RUC", "Address", "City", "Province", "Country",
             "Website", "Email", "Phone", "LinkedIn", "Twitter", "Notes", "Tags"]
    r_org = [[o.get(f, "") for f in ["num", "nombre", "industria", "ruc", "direccion", "ciudad",
                                       "provincia", "pais", "website", "email", "telefono",
                                       "linkedin", "twitter", "notas", "tags"]] for o in orgs]
    write_sheet(ws, h_org, r_org, [4, 24, 18, 14, 24, 14, 14, 10, 20, 26, 16, 20, 16, 30, 20])
    if orgs:
        sr = len(orgs) + 2
        ws.cell(row=sr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=9)
        ws.cell(row=sr, column=2, value=f"=COUNTA(B2:B{sr-1})").font = Font(name="Arial", bold=True, size=9)
    wb.save(OUT / "organizations.xlsx")
    print(f"  XLSX: {OUT / 'organizations.xlsx'}")

    # Contacts
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Contacts"
    h_ct = ["#", "Name", "Last Name", "Email", "Phone", "Role", "Organization", "Origin",
            "Type", "Status", "Priority", "Created", "Last Contact", "Products", "Notes", "Tags"]
    r_ct = []
    for c in contacts:
        oid = c.get("organization_id", "")
        org_name = org_idx[oid].get("nombre", "") if oid in org_idx else ""
        prods = resolve_products(c.get("product_ids", ""), prod_idx)
        r_ct.append([
            c.get("num", ""), c.get("nombre", ""), c.get("apellido", ""),
            c.get("email", ""), c.get("telefono_movil", ""), c.get("cargo", ""),
            org_name, c.get("origen", ""), c.get("tipo", ""), c.get("estado", ""),
            c.get("prioridad", ""), c.get("fecha_registro", ""),
            c.get("ultima_interaccion", ""), ", ".join(prods),
            c.get("notas", ""), c.get("tags", ""),
        ])
    write_sheet(ws2, h_ct, r_ct, [4, 14, 14, 26, 18, 18, 26, 14, 12, 10, 10, 16, 16, 30, 30, 20])
    if contacts:
        sr = len(contacts) + 2
        ws2.cell(row=sr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=9)
        ws2.cell(row=sr, column=2, value=f"=COUNTA(B2:B{sr-1})").font = Font(name="Arial", bold=True, size=9)
    wb2.save(OUT / "contacts.xlsx")
    print(f"  XLSX: {OUT / 'contacts.xlsx'}")

    # Products
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Products"
    h_pr = ["#", "Name", "Category", "Presentation", "Volume", "Price Ref.",
            "Coverage", "Fleet", "Regulatory", "Notes", "Tags"]
    r_pr = [[p.get(f, "") for f in ["num", "nombre", "categoria", "presentacion", "volumen",
                                      "precio_ref", "cobertura", "flota", "regulatorio", "notas", "tags"]] for p in products]
    write_sheet(ws3, h_pr, r_pr, [4, 30, 20, 20, 20, 14, 20, 8, 22, 30, 20])
    if products:
        sr = len(products) + 2
        ws3.cell(row=sr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=9)
        ws3.cell(row=sr, column=2, value=f"=COUNTA(B2:B{sr-1})").font = Font(name="Arial", bold=True, size=9)
    wb3.save(OUT / "products.xlsx")
    print(f"  XLSX: {OUT / 'products.xlsx'}")


# ── Run all ──

def run_all():
    OUT.mkdir(parents=True, exist_ok=True)
    orgs = load_csv(ORG_CSV)
    contacts = load_csv(CONTACT_CSV)
    products_list = load_csv(PRODUCT_CSV)
    org_idx = build_org_index(orgs)
    prod_idx = build_prod_index(products_list)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    print("\nExporting...")
    _pdf_orgs(orgs, ts)
    _pdf_contacts(contacts, org_idx, prod_idx, ts)
    _pdf_products(products_list, ts)
    _xlsx_all(orgs, contacts, org_idx, prod_idx, products_list, ts)
    print("  Done.\n")
    return True


if __name__ == "__main__":
    _ensure_deps()
    run_all()
